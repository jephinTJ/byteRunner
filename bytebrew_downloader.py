"""
Games DP1 Data - ByteBrew multi-game downloader

Reads games.xlsx and processes each Active=Yes row in Order.

Exact workflow per game:
    Open Game URL
    -> select Date Range
    -> Load saved funnel
    -> Close Load Filter popup
    -> select Build Version
    -> select GEO / Country
    -> Close Funnel Filters
    -> Apply
    -> Wait for Data Table
    -> Click Data Table three-dots
    -> Download CSV

Output:
    For Date Range = Yesterday:
        Games DP1 Data/YYYY-MM-DD/<Output Name>.csv

Notes:
- One persistent browser profile is reused for all games.
- First login may require manual login once.
- If one game fails, the script saves a debug screenshot and continues.
"""

from pathlib import Path
from datetime import date, datetime, timedelta
import re
import sys
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


BASE_URL = "https://dashboard.bytebrew.io/console"

APP_DIR = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).resolve().parent
SYSTEM_DIR = APP_DIR / "System Files"
UPLOAD_DIR = APP_DIR / "Upload files"
SYSTEM_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

CONFIG_FILE = SYSTEM_DIR / "games.xlsx" if (SYSTEM_DIR / "games.xlsx").exists() else APP_DIR / "games.xlsx"
PROFILE_DIR = SYSTEM_DIR / "bytebrew_profile"
DEBUG_DIR = SYSTEM_DIR / "debug"


LOG_CALLBACK = None
CURRENT_DATE_DIR = None  # Global tracker to ensure logs and screenshots route to the exact daily folder

def log(message):
    global CURRENT_DATE_DIR
    msg_str = f"[ByteBrew] {message}"
    print(msg_str, flush=True)
    if LOG_CALLBACK:
        try:
            LOG_CALLBACK(msg_str)
        except Exception:
            pass
            
    # Requirement 4: Append all terminal output to a master execution_log.txt in the daily folder
    if CURRENT_DATE_DIR:
        try:
            CURRENT_DATE_DIR.mkdir(parents=True, exist_ok=True)
            with open(CURRENT_DATE_DIR / "execution_log.txt", "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg_str}\n")
        except Exception:
            pass


def safe_name(value):
    value = str(value or "").strip()
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    return value or "Game"


def screenshot(page, game, name):
    global CURRENT_DATE_DIR
    # Requirement 2: Route all screenshots directly to the daily output folder instead of the generic debug folder
    target_dir = CURRENT_DATE_DIR if CURRENT_DATE_DIR else DEBUG_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = target_dir / f"{safe_name(game)}_{name}_{stamp}.png"
    try:
        page.screenshot(path=str(path), full_page=True)
        log(f"Debug screenshot: {path.name}")
    except Exception:
        pass
    return path


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_active(value):
    return cell_text(value).lower() in {"yes", "y", "true", "1", "active"}


def load_games():
    if not CONFIG_FILE.exists():
        raise FileNotFoundError(
            f"Could not find {CONFIG_FILE.name} next to the Python script."
        )

    workbook = load_workbook(CONFIG_FILE, data_only=True)
    sheet_name = "Dp1" if "Dp1" in workbook.sheetnames else "Games"
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f'games.xlsx must contain a sheet named "{sheet_name}".')

    ws = workbook[sheet_name]
    headers = [cell_text(c.value) for c in ws[1]]
    required = {
        "Order", "Active", "Game Name", "Game ID", "Page", "Saved Funnel 1",
        "Saved Funnel 2", "Date Range", "Output Name"
    }
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(
            "games.xlsx is missing required column(s): " + ", ".join(missing)
        )

    games = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        if not is_active(data.get("Active")):
            continue

        game_id = cell_text(data.get("Game ID"))
        game_name = cell_text(data.get("Game Name"))

        if not game_id:
            log(f'Skipping "{game_name or "Unnamed game"}": Game ID is blank.')
            continue

        try:
            order = int(data.get("Order") or 9999)
        except Exception:
            order = 9999

        games.append({
            "order": order,
            "game_name": game_name or game_id,
            "game_id": game_id,
            "page": cell_text(data.get("Page")) or "funnelexplorer",
            "saved_funnels": [
                f for f in [
                    cell_text(data.get("Saved Funnel 1")),
                    cell_text(data.get("Saved Funnel 2")),
                ] if f
            ],
            "date_range": cell_text(data.get("Date Range")) or "Yesterday",
            "build_version": cell_text(data.get("Build Version")),
            "country": cell_text(data.get("Country")),
            "output_name": cell_text(data.get("Output Name")) or game_name or game_id,
        })

    games.sort(key=lambda x: x["order"])

    if not games:
        raise RuntimeError(
            "No active games found. Set Active = Yes for at least one row in games.xlsx."
        )

    return games


def output_date_for_preset(preset):
    p = str(preset or "").strip().lower()
    today = date.today()

    if p in {"yesterday", ""}:
        return today - timedelta(days=1)
    if p == "today":
        return today

    try:
        return datetime.strptime(p, "%Y-%m-%d").date()
    except Exception:
        return today - timedelta(days=1)


def click_text(page, text, exact=True, timeout=15000):
    candidates = [
        page.get_by_text(text, exact=exact),
        page.get_by_role("button", name=text, exact=exact),
    ]

    last_error = None
    for locator in candidates:
        try:
            locator.first.wait_for(state="visible", timeout=timeout)
            locator.first.click()
            return
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f'Could not click "{text}".') from last_error


def click_visible_exact_text(page, text, timeout=15000):
    locator = page.get_by_text(text, exact=True)

    end_time = datetime.now().timestamp() + timeout / 1000
    while datetime.now().timestamp() < end_time:
        for i in range(locator.count()):
            try:
                item = locator.nth(i)
                if item.is_visible():
                    item.click()
                    return item
            except Exception:
                pass
        page.wait_for_timeout(250)

    raise RuntimeError(f'Could not find visible exact value "{text}".')


def wait_for_login_if_needed(page, target_url, email=None, password=None):
    page.wait_for_timeout(2500)

    if "/console/" in page.url:
        return

    log("ByteBrew login is required.")

    if not email or not password:
        raise RuntimeError("LOGIN_REQUIRED: Credentials are empty.")

    try:
        log("Attempting automated credential login...")
        email_field = page.locator('input[type="email"], input[name="email"], input[placeholder*="email" i]').first
        email_field.wait_for(state="visible", timeout=8000)
        email_field.fill(email)

        pass_field = page.locator('input[type="password"], input[name="password"]').first
        pass_field.fill(password)

        submit_btn = page.locator('button[type="submit"], button:has-text("Log In"), button:has-text("Sign In")').first
        submit_btn.click()

        page.wait_for_url("**/console/**", timeout=15000)
        log("Automated login successful.")
    except Exception as err:
        log(f"Automated login failed: {err}")
        raise RuntimeError(f"LOGIN_FAILED: {err}")

    page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(2500)


def open_game(page, game, email=None, password=None):
    url = f'{BASE_URL}/{game["game_id"]}/{game["page"]}'
    log(f'Opening {game["game_name"]}: {url}')

    # Bulletproof Double-Tap: Try loading up to 3 times to survive ByteBrew server chokes
    for attempt in range(1, 4):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            break
        except Exception as e:
            if attempt == 3:
                raise RuntimeError(f"Server dead. Failed to load URL after 3 attempts: {e}")
            log(f"Server choke detected. Retrying URL load ({attempt}/3)...")
            page.wait_for_timeout(3000)

    wait_for_login_if_needed(page, url, email=email, password=password)

    if game["game_id"] not in page.url or game["page"] not in page.url:
        for attempt in range(1, 4):
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                break
            except Exception as e:
                if attempt == 3:
                    raise RuntimeError(f"Server dead on redirect. Failed to load URL: {e}")
                log(f"Server choke on redirect. Retrying URL load ({attempt}/3)...")
                page.wait_for_timeout(3000)

    page.wait_for_timeout(1500)
    log(f'{game["game_name"]} Funnel Explorer opened.')


def set_date_range(page, preset, game_name):
    target_dt = output_date_for_preset(preset)
    target_iso = target_dt.isoformat()
    target_year = target_dt.year
    target_month_num = target_dt.month
    target_month_short = target_dt.strftime("%b")
    target_month_full = target_dt.strftime("%B")
    target_day = target_dt.day

    for attempt in range(1, 4):
        log(f"[Clicker] Setting Top Date Range to: '{target_iso}' (Attempt {attempt})...")
        
        if attempt > 1:
            # Click neutral background space to reset state instead of Escape (which cancels)
            page.mouse.click(10, 10)
            page.wait_for_timeout(400)

        # 1. Open the main date range dropdown button using the exact ID you found
        date_control = page.locator('#daterange-btn, button:has-text(" - "), [role="button"]:has-text(" - ")')
        clicked = False
        for i in range(date_control.count()):
            try:
                el = date_control.nth(i)
                if el.is_visible() and ("202" in el.inner_text() or " - " in el.inner_text()):
                    el.click()
                    clicked = True
                    break
            except Exception:
                pass

        if not clicked:
            candidates = page.get_by_text(re.compile(
                r"(January|February|March|April|May|June|July|August|September|October|November|December).*\d{4}\s*-\s*", re.I
            ))
            for i in range(candidates.count()):
                try:
                    el = candidates.nth(i)
                    if el.is_visible():
                        el.click()
                        clicked = True
                        break
                except Exception:
                    pass

        if not clicked:
            screenshot(page, game_name, "date_control_not_found")
            raise RuntimeError("Could not locate ByteBrew top date-range control.")

        page.wait_for_timeout(400)

        # 2. Click "Custom Range" to expand calendar view
        custom_btn = page.get_by_text("Custom Range", exact=True)
        if custom_btn.count() == 0:
            custom_btn = page.locator('li:has-text("Custom Range"), div:has-text("Custom Range"), button:has-text("Custom Range")')
        
        clicked_custom = False
        # Loop to explicitly find the VISIBLE "Custom Range" and ignore ByteBrew's hidden ghost nodes
        for i in range(custom_btn.count()):
            try:
                el = custom_btn.nth(i)
                if el.is_visible():
                    el.click()
                    clicked_custom = True
                    break
            except Exception:
                pass
                
        if not clicked_custom:
            try:
                custom_btn.first.click(force=True)
            except Exception:
                pass

        page.wait_for_timeout(500)

        # 3. Handle Month Navigation (◄ / ►) in the Top Date Calendar widget
        def get_visible_calendar_months():
            month_headers = page.locator('.daterangepicker .month, .drp-calendar .month, th.month')
            months_found = []
            month_map = {"Jan":1, "Feb":2, "Mar":3, "Apr":4, "May":5, "Jun":6, "Jul":7, "Aug":8, "Sep":9, "Oct":10, "Nov":11, "Dec":12}
            for idx in range(month_headers.count()):
                try:
                    txt = month_headers.nth(idx).inner_text().strip()
                    if txt:
                        # Bulletproof regex handles "August 2026", "AUGUST2026", "Aug\n2026", etc. safely
                        m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*(\d{4})', txt, re.I)
                        if m:
                            y_num = int(m.group(2))
                            m_num = month_map[m.group(1).title()]
                            months_found.append((y_num, m_num))
                except Exception:
                    pass
            return months_found

        target_month_idx = target_year * 12 + target_month_num
        for _ in range(24):
            visible_months = get_visible_calendar_months()
            if not visible_months:
                break
            
            month_indices = [y * 12 + m for y, m in visible_months]
            if target_month_idx in month_indices:
                break
            
            min_visible = min(month_indices)
            max_visible = max(month_indices)
            
            if target_month_idx < min_visible:
                prev_arrow = page.locator('.daterangepicker th.prev, .drp-calendar th.prev, .prev.available, th:has(.fa-chevron-left), .fa-chevron-left')
                if prev_arrow.count() and prev_arrow.first.is_visible():
                    prev_arrow.first.click()
                else:
                    page.keyboard.press("ArrowLeft")
                page.wait_for_timeout(300)
            elif target_month_idx > max_visible:
                next_arrow = page.locator('.daterangepicker th.next, .drp-calendar th.next, .next.available, th:has(.fa-chevron-right), .fa-chevron-right')
                if next_arrow.count() and next_arrow.first.is_visible():
                    next_arrow.first.click()
                else:
                    page.keyboard.press("ArrowRight")
                page.wait_for_timeout(300)

        # 4. Click Day Cell twice (Start Date & End Date for single-day range)
        day_clicked = False
        calendars = page.locator('.daterangepicker .drp-calendar, .daterangepicker .calendar-table, .calendar.left, .calendar.right')
        for c_idx in range(calendars.count()):
            try:
                cal = calendars.nth(c_idx)
                if not cal.is_visible():
                    continue
                header_text = cal.locator('.month, th.month').inner_text().strip()
                if (target_month_short in header_text or target_month_full in header_text) and str(target_year) in header_text:
                    day_cells = cal.locator(f'td.available:not(.off):text-is("{target_day}"), td:not(.off):text-is("{target_day}")')
                    if day_cells.count() > 0 and day_cells.first.is_visible():
                        day_cells.first.click()
                        page.wait_for_timeout(200)
                        day_cells.first.click()
                        day_clicked = True
                        break
            except Exception:
                pass

        if not day_clicked:
            # Tightened fallback to strictly block ".off" ghost days from adjacent months
            all_day_cells = page.locator(f'.daterangepicker td.available:not(.off):text-is("{target_day}")')
            for i in range(all_day_cells.count()):
                try:
                    cell = all_day_cells.nth(i)
                    if cell.is_visible():
                        cell.click()
                        page.wait_for_timeout(200)
                        cell.click()
                        day_clicked = True
                        break
                except Exception:
                    pass

        if not day_clicked:
            screenshot(page, game_name, "custom_date_day_click_failed")
            raise RuntimeError(f"Could not click day {target_day} in Custom Range calendar.")

        page.wait_for_timeout(300)

        # 5. Click "Apply" in DateRangePicker if visible
        apply_btn = page.locator('.daterangepicker .applyBtn, .daterangepicker button:has-text("Apply"), button.applyBtn')
        if apply_btn.count() > 0 and apply_btn.first.is_visible():
            apply_btn.first.click()
            page.wait_for_timeout(500)
        else:
            # Top Date picker doesn't have an apply button. Click empty background space to commit the date and close the popup.
            page.mouse.click(10, 10)
            page.wait_for_timeout(500)

        log(f"[Checker] Verifying Top Date Range updated to '{target_iso}'...")
        try:
            # Target the absolute hardcoded ID from your inspection
            active_date_btn = page.locator('#daterange-btn, button:has-text(" - "), [role="button"]:has-text(" - ")').first
            
            # Poll the exact DOM element for up to 5 seconds to give ByteBrew's UI time to update the text
            success = False
            btn_text = ""
            for _ in range(10):
                btn_text = active_date_btn.inner_text().strip()
                has_year = str(target_year) in btn_text
                has_day = bool(re.search(rf"\b{target_day}\b", btn_text))
                has_month = (target_month_short in btn_text) or (target_month_full in btn_text)
                
                if has_year and has_day and has_month:
                    success = True
                    break
                page.wait_for_timeout(500)
                
            if success:
                log(f"[Checker] Success: Top Date Range visually confirmed as '{btn_text}'.")
                return
            else:
                log(f"[Checker] Mismatch: UI shows '{btn_text}'. Retrying date selection...")
        except Exception as e:
            log(f"[Checker] Error while reading Top Date text ({e}). Retrying...")

    screenshot(page, game_name, "top_date_verification_failed")
    raise RuntimeError(f"[Checker] FATAL: Failed to verify Top Date Range set to '{target_iso}' after 3 attempts.")


def load_saved_funnel(page, saved_funnel, game_name):
    """
    Load a saved ByteBrew filter for either Funnel Explorer or Mechanics.
    Includes QQ2 and QQ4 State Verification with API wipe self-healing.
    """
    if not saved_funnel:
        raise RuntimeError("Saved Filter is blank in games.xlsx.")

    for load_attempt in range(1, 5):
        log(f"[Checker] Verifying if preset '{saved_funnel}' AND 'Countries' pill are active (Attempt {load_attempt})...")
        
        preset_ok = False
        pill_ok = False
        try:
            current_preset = page.locator("#loaded-filter-name").inner_text(timeout=3000).strip()
            if current_preset == saved_funnel:
                preset_ok = True
                # Wait up to 4 seconds for the API payload to physically render the 'Countries' tag
                page.locator("#filter-result-area-1 span.tag:has-text('Countries')").wait_for(state="visible", timeout=4000)
                pill_ok = True
        except Exception:
            pass
            
        if preset_ok and pill_ok:
            log(f"[Checker] Success: '{saved_funnel}' and 'Countries' pill are securely locked in the DOM.")
            return

        # If we reach here, either the preset text mismatched or the Geo pill vanished
        if load_attempt > 1:
            if preset_ok and not pill_ok:
                log(f"[Clicker] ByteBrew API glitch detected (Missing Geo). Force-reloading '{saved_funnel}'...")
            else:
                log(f"[Clicker] Preset mismatch. Force-reloading '{saved_funnel}'...")
            page.wait_for_timeout(1000)
        else:
            log(f"[Clicker] Loading saved filter: '{saved_funnel}'")

        click_text(page, "Load Filter")
        page.wait_for_timeout(1000)  # Wait for popup animation

        # ByteBrew uses different modal titles for different explorer pages.
        modal_title = page.get_by_text(
            re.compile(r"Load\s+(Funnel|Mechanic)\s+Filter", re.I)
        ).first

        try:
            modal_title.wait_for(state="visible", timeout=15000)
        except PlaywrightTimeoutError:
            screenshot(page, game_name, "load_filter_modal_not_found")
            raise RuntimeError(
                'Could not find "Load Funnel Filter" or "Load Mechanic Filter" popup.'
            )

        log(f'Opened popup: "{modal_title.inner_text().strip() if modal_title else "Load Filter"}"')

        selected = False
        last_scroll_signature = None

        for attempt in range(1, 31):
            # Prefer exact row text so PartA never accidentally matches PartB.
            row = page.locator("tr").filter(
                has=page.get_by_text(saved_funnel, exact=True)
            )

            if row.count() > 0:
                for r in range(row.count()):
                    try:
                        target_row = row.nth(r)
                        target_row.scroll_into_view_if_needed()
                        select_button = target_row.get_by_text("Select", exact=True)

                        if select_button.count():
                            select_button.first.click()
                            selected = True
                            break
                    except Exception:
                        pass

            if selected:
                break

            # Fallback for non-table rendering.
            exact_name = page.get_by_text(saved_funnel, exact=True)

            for i in range(exact_name.count()):
                try:
                    item = exact_name.nth(i)
                    if not item.is_visible():
                        continue

                    item.scroll_into_view_if_needed()

                    for levels in range(1, 8):
                        container = item.locator("xpath=" + "/.." * levels)
                        button = container.get_by_text("Select", exact=True)

                        if button.count() and button.first.is_visible():
                            button.first.click()
                            selected = True
                            break

                    if selected:
                        break
                except Exception:
                    pass

            if selected:
                break

            # Scroll the currently visible Load Filter modal.
            try:
                # Added 'r' prefix to fix Python \s SyntaxWarning
                signature = page.evaluate(
                    r"""
                    () => {
                        const title = [...document.querySelectorAll('*')]
                            .find(el =>
                                /Load\s+(Funnel|Mechanic)\s+Filter/i.test(
                                    el.textContent?.trim() || ''
                                )
                            );

                        if (!title) return 'no-title';

                        let modal = title.closest('.modal, [role="dialog"]');
                        if (!modal) modal = title.parentElement;

                        const all = [modal, ...modal.querySelectorAll('*')];
                        let moved = 0;
                        let signature = [];

                        for (const el of all) {
                            if (!el) continue;

                            const style = getComputedStyle(el);
                            const canScroll =
                                (
                                    style.overflowY === 'auto' ||
                                    style.overflowY === 'scroll'
                                ) &&
                                el.scrollHeight > el.clientHeight + 5;

                            if (canScroll) {
                                const before = el.scrollTop;

                                el.scrollTop = Math.min(
                                    el.scrollTop +
                                        Math.max(350, el.clientHeight * 0.8),
                                    el.scrollHeight
                                );

                                if (el.scrollTop !== before) moved++;

                                signature.push(
                                    `${Math.round(el.scrollTop)}/` +
                                    `${Math.round(el.scrollHeight)}`
                                );
                            }
                        }

                        return `${moved}|${signature.join(',')}`;
                    }
                    """
                )

                # Pagination fallback.
                if signature == last_scroll_signature or signature.startswith("0|"):
                    next_buttons = page.get_by_text("Next", exact=True)

                    for i in range(next_buttons.count()):
                        try:
                            nxt = next_buttons.nth(i)
                            if nxt.is_visible() and nxt.is_enabled():
                                nxt.click()
                                page.wait_for_timeout(500)
                                break
                        except Exception:
                            pass

                last_scroll_signature = signature

            except Exception:
                page.keyboard.press("PageDown")

            page.wait_for_timeout(300)

        if not selected:
            screenshot(
                page,
                game_name,
                f"saved_filter_{safe_name(saved_funnel)}_not_found"
            )
            raise RuntimeError(
                f'Saved filter "{saved_funnel}" was not found after scrolling.'
            )

        log(f"[Clicker] Clicked '{saved_funnel}'. Waiting for popup to auto-close...")
        
        try:
            modal_title.wait_for(state="hidden", timeout=6000)
            log("[Checker] Load Filter popup auto-closed.")
        except Exception:
            log("[Clicker] Popup did not auto-close. Forcing Escape...")
            page.keyboard.press("Escape")
            
        # Give DOM a moment to fetch the corrupted/clean payload before next loop iteration verifies it
        page.wait_for_timeout(2000)

    screenshot(page, game_name, "preset_api_wipe_failed")
    raise RuntimeError(f"[Checker] FATAL: ByteBrew API repeatedly wiped the '{saved_funnel}' payload after 4 attempts.")


def select_filter_value(page, category, value, game_name):
    if not value:
        return

    log(f"Ensuring {category} = {value} is selected...")

    # Left-side filter category.
    click_text(page, category, exact=False)
    page.wait_for_timeout(400)

    # Safe toggle check: Only click if NOT already selected/checked
    try:
        was_checked = page.evaluate(f'''(val) => {{
            const items = [...document.querySelectorAll('tr, li, .list-group-item, div, label')];
            const matching = items.filter(el => el.innerText?.trim().toUpperCase() === val.toUpperCase());
            if (!matching.length) return false;

            const target = matching[0];
            const checkbox = target.querySelector('input[type="checkbox"]');
            if (checkbox) {{
                if (checkbox.checked) return true; // Already checked, keep active
                checkbox.click();
                return true;
            }}

            const cls = (target.className || '') + ' ' + (target.getAttribute('class') || '');
            if (/active|selected|checked/i.test(cls)) {{
                return true; // Already active, do not click
            }}

            target.click();
            return true;
        }}''', value)

        if not was_checked:
            target = page.get_by_text(value, exact=True)
            for i in range(target.count()):
                item = target.nth(i)
                if item.is_visible():
                    item.click()
                    break
    except Exception as err:
        log(f"Notice during {category}={value} selection: {err}")

    page.wait_for_timeout(350)


def ensure_graph_type_bar(page, game_name=None):
    """Silently forces Graph Type = Bar using safe DOM checks."""
    try:
        page.evaluate(r'''() => {
            try {
                const elements = Array.from(document.querySelectorAll('*'));
                const barBtns = elements.filter(el => (el.innerText || '').trim().toLowerCase() === 'bar');
                for (let btn of barBtns) {
                    let p = btn.parentElement;
                    if (p && (p.innerText.toLowerCase().includes('line') || p.innerText.toLowerCase().includes('stacked'))) {
                        const cls = (btn.className || '').toLowerCase() + ' ' + (p.className || '').toLowerCase();
                        if (!cls.includes('active') && !cls.includes('selected')) {
                            btn.click();
                        }
                        break;
                    }
                }
            } catch(e) {}
        }''')
        page.wait_for_timeout(300)
    except Exception:
        pass


def configure_build_and_geo(page, build_version, country, game_name):
    if not build_version and not country:
        log("No Build Version or Country filter configured; skipping Funnel Filters.")
        return

    log("Opening Funnel Filters...")
    click_text(page, "Funnel Filters")

    page.get_by_text("Event Filters", exact=False).first.wait_for(
        state="visible", timeout=15000
    )

    if build_version:
        select_filter_value(page, "BUILD VERSION", build_version, game_name)

    if country:
        select_filter_value(page, "GEO", country, game_name)

    # Close Event Filters only after BOTH Build and GEO are selected.
    log("Closing Event Filters...")
    close_candidates = [
        page.get_by_role("button", name=re.compile("close", re.I)),
        page.locator('[aria-label*="close" i]'),
        page.locator('[title*="close" i]'),
        page.locator("text=×"),
    ]

    closed = False
    for candidate in close_candidates:
        try:
            if candidate.count():
                for i in range(candidate.count() - 1, -1, -1):
                    el = candidate.nth(i)
                    if el.is_visible():
                        el.click()
                        closed = True
                        break
            if closed:
                break
        except Exception:
            pass

    if not closed:
        page.keyboard.press("Escape")

    page.get_by_text("Funnel Graph", exact=True).first.wait_for(
        state="visible", timeout=15000
    )
    log("Build and GEO filters selected.")


def save_filter(page, saved_funnel, game_name):
    """
    Click the main-page Save Filter control BEFORE Apply.

    ByteBrew installations may either save immediately or open a confirmation
    dialog. If a dialog appears, this function tries Update/Save/Confirm.
    """
    log("Saving filter before Apply...")

    save_link = page.get_by_text(re.compile(r"^\+?\s*Save Filter\s*$", re.I))
    clicked = False

    for i in range(save_link.count()):
        try:
            el = save_link.nth(i)
            if el.is_visible():
                el.click()
                clicked = True
                break
        except Exception:
            pass

    if not clicked:
        # Role/button fallback.
        button = page.get_by_role("button", name=re.compile(r"Save Filter", re.I))
        if button.count():
            button.first.click()
            clicked = True

    if not clicked:
        screenshot(page, game_name, "save_filter_button_not_found")
        raise RuntimeError('Could not locate "Save Filter".')

    page.wait_for_timeout(700)

    # Handle an optional save/update modal without assuming one always exists.
    dialogs = page.locator('[role="dialog"]')
    visible_dialog = None

    for i in range(dialogs.count()):
        try:
            if dialogs.nth(i).is_visible():
                visible_dialog = dialogs.nth(i)
                break
        except Exception:
            pass

    if visible_dialog is not None:
        # If there is an empty text field, preserve the configured saved funnel name.
        try:
            inputs = visible_dialog.locator('input[type="text"], input:not([type])')
            for i in range(inputs.count()):
                inp = inputs.nth(i)
                if inp.is_visible() and not (inp.input_value() or "").strip():
                    inp.fill(saved_funnel)
                    break
        except Exception:
            pass

        confirmed = False
        for pattern in [r"^Update$", r"^Save$", r"^Confirm$", r"^OK$"]:
            try:
                btn = visible_dialog.get_by_role(
                    "button", name=re.compile(pattern, re.I)
                )
                if btn.count() and btn.first.is_visible():
                    btn.first.click()
                    confirmed = True
                    break
            except Exception:
                pass

        if confirmed:
            try:
                visible_dialog.wait_for(state="hidden", timeout=8000)
            except Exception:
                pass
        else:
            screenshot(page, game_name, "save_filter_dialog_unhandled")
            raise RuntimeError(
                "Save Filter opened a confirmation dialog that the script "
                "could not identify. See the debug screenshot."
            )

    log("Filter saved.")



def verify_filter_pills(page, game_name):
    """Checker function to explicitly look for 'Countries' in the #filter-result-area-1 box. (QQ4)"""
    log("[Checker] Verifying active filter pills (QQ4) against API wipe...")
    try:
        pill = page.locator("#filter-result-area-1 span.tag:has-text('Countries')")
        pill.wait_for(state="visible", timeout=4000)
        log("[Checker] Success: 'Countries' filter pill is securely locked in the DOM.")
    except PlaywrightTimeoutError:
        screenshot(page, game_name, "filter_pill_wiped")
        raise RuntimeError("[Checker] FATAL: ByteBrew background API refresh wiped the 'Countries' filter! Aborting to prevent dirty data.")


def ensure_graph_type_bar(page):
    """Silently forces Graph Type = Bar using raw DOM strikes to avoid multi-row exports."""
    log("[Clicker] Checking Graph Type...")
    try:
        bar_radio = page.locator("#dash-graph-bar")
        if bar_radio.count() > 0:
            # Check the true HTML state of the hidden radio button
            if not bar_radio.first.is_checked():
                log("[Clicker] Bar chart not active. Clicking Bar toggle...")
                page.locator("label#barToggle").first.click(force=True)
                page.wait_for_timeout(400)
            
            log("[Checker] Verifying Graph Type is securely set to 'Bar'...")
            if bar_radio.first.is_checked():
                log("[Checker] Success: Graph Type is Bar.")
    except Exception as e:
        log(f"[Checker] Notice during Graph Type check: {e}")


def configure_install_date_filter(page, preset, game_name):
    """
    Mechanics-only install-date filter.

    Flow:
      Filters -> Segment -> Settings
      -> USER INSTALL DATE / USER INSTALL DATE (*)
      -> set Start = target_date
      -> set End = target_date
      -> Save
    """
    target_date = output_date_for_preset(preset).isoformat()
    log(f"Setting User Install Date = {target_date}")

    # ------------------------------------------------------------
    # 1. Click Filters -> Segment -> Settings
    # ------------------------------------------------------------
    filters_heading = page.get_by_text("Filters", exact=True).first

    try:
        filters_heading.wait_for(state="visible", timeout=10000)
        filters_box = filters_heading.bounding_box()
    except Exception:
        filters_box = None

    if not filters_box:
        screenshot(page, game_name, "filters_heading_not_found")
        raise RuntimeError(
            'Could not locate the main "Filters" section.'
        )

    settings_candidates = page.get_by_text("Settings", exact=True)
    nearby_settings = []

    filters_x = filters_box["x"]
    filters_y = filters_box["y"]

    for i in range(settings_candidates.count()):
        try:
            item = settings_candidates.nth(i)

            if not item.is_visible():
                continue

            box = item.bounding_box()
            if not box:
                continue

            in_filters_column = box["x"] >= filters_x - 30
            below_heading = box["y"] >= filters_y
            reasonably_close = box["y"] <= filters_y + 450

            if in_filters_column and below_heading and reasonably_close:
                distance = (
                    abs(box["x"] - filters_x)
                    + abs(box["y"] - filters_y)
                )
                nearby_settings.append((distance, item))

        except Exception:
            pass

    if not nearby_settings:
        screenshot(page, game_name, "segment_settings_not_found")
        raise RuntimeError(
            'Could not find Filters -> Segment -> Settings.'
        )

    nearby_settings.sort(key=lambda x: x[0])
    segment_settings = nearby_settings[0][1]

    log("Clicking Filters → Segment → Settings...")
    segment_settings.click()

    # ------------------------------------------------------------
    # 2. Wait for Segment Settings modal
    # ------------------------------------------------------------
    modal_title = page.get_by_text("Segment Settings", exact=True).first

    try:
        modal_title.wait_for(state="visible", timeout=10000)
    except PlaywrightTimeoutError:
        screenshot(page, game_name, "segment_settings_modal_not_open")
        raise RuntimeError(
            'Clicking Segment Settings did not open the popup.'
        )

    log("Segment Settings opened.")

    # ------------------------------------------------------------
    # 3. Select USER INSTALL DATE
    # ------------------------------------------------------------
    log('Selecting "USER INSTALL DATE"...')

    # Added 'r' prefix to fix Python \s SyntaxWarning
    clicked = page.evaluate(
        r"""
        () => {
            const norm = (s) =>
                (s || '')
                    .replace(/\s+/g, ' ')
                    .trim()
                    .toUpperCase();

            const visible = (el) => {
                const r = el.getBoundingClientRect();
                const style = getComputedStyle(el);

                return (
                    r.width > 0 &&
                    r.height > 0 &&
                    style.display !== 'none' &&
                    style.visibility !== 'hidden'
                );
            };

            const candidates = [...document.querySelectorAll('*')]
                .filter(el => {
                    if (!visible(el)) return false;

                    const txt = norm(el.innerText);

                    return (
                        txt === 'USER INSTALL DATE' ||
                        txt === 'USER INSTALL DATE (*)'
                    );
                });

            if (!candidates.length) return false;

            // Click the smallest exact-text node.
            candidates.sort((a, b) => {
                const ra = a.getBoundingClientRect();
                const rb = b.getBoundingClientRect();

                return (ra.width * ra.height) - (rb.width * rb.height);
            });

            candidates[0].click();
            return true;
        }
        """
    )

    if not clicked:
        screenshot(page, game_name, "user_install_date_option_not_found")
        raise RuntimeError(
            'Could not click "USER INSTALL DATE" in Segment Settings.'
        )

    log('"USER INSTALL DATE" selected.')

    # ------------------------------------------------------------
    # 4. Wait for existing Start / End date fields
    # ------------------------------------------------------------
    page.get_by_text(
        re.compile(r"Filter by User Install Date", re.I)
    ).first.wait_for(state="visible", timeout=10000)

    page.get_by_text(
        re.compile(r"Install Date Range Picker", re.I)
    ).first.wait_for(state="visible", timeout=10000)

    log("Looking for Start and End date fields...")

    date_inputs = []

    for _ in range(30):
        date_inputs = []

        inputs = page.locator("input")

        for i in range(inputs.count()):
            try:
                inp = inputs.nth(i)

                if not inp.is_visible():
                    continue

                input_type = (inp.get_attribute("type") or "").lower()
                value = (inp.input_value() or "").strip()
                placeholder = (inp.get_attribute("placeholder") or "").strip()

                looks_like_date = (
                    input_type == "date"
                    or re.fullmatch(r"\d{4}-\d{2}-\d{2}", value or "")
                    or re.search(r"date", placeholder, re.I)
                )

                if looks_like_date:
                    date_inputs.append(inp)

            except Exception:
                pass

        if len(date_inputs) >= 2:
            break

        page.wait_for_timeout(250)

    if len(date_inputs) < 2:
        screenshot(page, game_name, "install_date_inputs_not_found")
        raise RuntimeError(
            "Install Date Range Picker is enabled, but Start/End date inputs "
            "could not be found."
        )

    # ------------------------------------------------------------
    # 5. Select Start and End from ByteBrew's calendar.
    #
    # The inputs are READONLY (jQuery datepicker), so .fill() cannot work.
    # Click each input, then click the target day in the visible calendar.
    # ------------------------------------------------------------
    from datetime import datetime as _dt

    target_dt = _dt.strptime(target_date, "%Y-%m-%d")
    target_year = target_dt.year
    target_month_zero_based = target_dt.month - 1
    target_day = target_dt.day

    def choose_date_from_picker(input_locator, label):
        log(f"[Clicker] Selecting {label} date = {target_date}")

        input_locator.click()
        page.wait_for_timeout(500)

        # jQuery UI datepicker is indicated by ByteBrew's:
        #   class="fc-datepicker hasDatepicker"
        selected = False

        for _ in range(24):
            # STRICT DOM LOCK: Target exact year and zero-indexed month attributes.
            day_links = page.locator(
                f'td[data-handler="selectDay"]'
                f'[data-year="{target_year}"]'
                f'[data-month="{target_month_zero_based}"] a'
            )

            for i in range(day_links.count()):
                try:
                    day_link = day_links.nth(i)
                    if (
                        day_link.is_visible()
                        and day_link.inner_text().strip() == str(target_day)
                    ):
                        day_link.click()
                        selected = True
                        break
                except Exception:
                    pass

            if selected:
                break

            # If the required month is not currently displayed, inspect the
            # visible datepicker title and navigate Prev/Next.
            title = page.locator(".ui-datepicker-title:visible")
            current_year = None
            current_month = None

            if title.count():
                try:
                    title_text = title.first.inner_text().strip()
                    # Bulletproof regex for User Install Date calendar headers (handles missing spaces)
                    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*(\d{4})', title_text, re.I)
                    if m:
                        month_map = {"Jan":1, "Feb":2, "Mar":3, "Apr":4, "May":5, "Jun":6, "Jul":7, "Aug":8, "Sep":9, "Oct":10, "Nov":11, "Dec":12}
                        current_month = month_map[m.group(1).title()]
                        current_year = int(m.group(2))
                except Exception:
                    pass

            if current_year is not None and current_month is not None:
                current_index = current_year * 12 + current_month
                target_index = target_year * 12 + target_dt.month

                if current_index < target_index:
                    next_btn = page.locator(".ui-datepicker-next")
                    if next_btn.count():
                        # Evaluate raw JS click to bypass Playwright's strict visibility blocks on styled pseudo-elements
                        next_btn.first.evaluate("el => el.click()")
                        page.wait_for_timeout(300)
                        continue

                elif current_index > target_index:
                    prev_btn = page.locator(".ui-datepicker-prev")
                    if prev_btn.count():
                        # Evaluate raw JS click to bypass Playwright's strict visibility blocks on styled pseudo-elements
                        prev_btn.first.evaluate("el => el.click()")
                        page.wait_for_timeout(300)
                        continue

            page.wait_for_timeout(200)

        if not selected:
            screenshot(page, game_name, f"install_date_{label.lower()}_calendar_failed")
            raise RuntimeError(f"[Clicker] Could not select {target_date} from the {label} calendar.")

        log(f"[Checker] Verifying {label} date input value matches '{target_date}'...")
        # Wait for ByteBrew to update the readonly input.
        for _ in range(20):
            try:
                if input_locator.input_value().strip() == target_date:
                    return
            except Exception:
                pass

            page.wait_for_timeout(150)

        actual = ""
        try:
            actual = input_locator.input_value()
        except Exception:
            pass

        raise RuntimeError(
            f"{label} date calendar click completed, but input value is "
            f'"{actual}" instead of "{target_date}".'
        )

    choose_date_from_picker(date_inputs[0], "Start")
    choose_date_from_picker(date_inputs[1], "End")

    start_value = date_inputs[0].input_value().strip()
    end_value = date_inputs[1].input_value().strip()

    if start_value != target_date or end_value != target_date:
        screenshot(page, game_name, "install_date_selection_failed")
        raise RuntimeError(
            f"Install Date values did not update correctly. "
            f"Start={start_value}, End={end_value}, expected={target_date}"
        )

    log(f"User Install Date set: {target_date} → {target_date}")

    # ------------------------------------------------------------
    # 6. Click green Save in Segment Settings
    # ------------------------------------------------------------
    log("[Checker] Verifying exact start/end input matches before saving...")
    log("[Clicker] Saving Segment Settings...")

    save_candidates = page.get_by_text("Save", exact=True)
    visible_saves = []

    for i in range(save_candidates.count()):
        try:
            btn = save_candidates.nth(i)

            if not btn.is_visible():
                continue

            box = btn.bounding_box()
            if box:
                visible_saves.append(
                    (box["x"] + box["y"], btn)
                )
        except Exception:
            pass

    if not visible_saves:
        screenshot(page, game_name, "segment_settings_save_not_found")
        raise RuntimeError(
            'Could not find the Segment Settings "Save" button.'
        )

    # Bottom/right-most Save button in the open popup.
    visible_saves.sort(key=lambda x: x[0], reverse=True)
    visible_saves[0][1].click()

    try:
        modal_title.wait_for(state="hidden", timeout=8000)
    except Exception:
        pass

    # Strict wait to allow the invisible .modal-backdrop CSS animation to fully clear the screen
    page.wait_for_timeout(1500)
    log("Segment Settings saved.")


def apply_funnel(page, game_name):
    # Safety guard in case ByteBrew's Save Filter modal is somehow open.
    save_modal = page.locator("#filtersavemodal")
    try:
        if save_modal.count() and save_modal.is_visible():
            log("Unexpected Save Filter modal detected; closing it.")
            page.keyboard.press("Escape")
            save_modal.wait_for(state="hidden", timeout=5000)
    except Exception:
        pass

    # Activate dormant function: Force the UI to Bar Chart so empty datasets render a 0-value table instead of timing out
    ensure_graph_type_bar(page)

    log("[Clicker] Clicking Apply...")

    apply_button = page.get_by_role("button", name="Apply", exact=True)
    if apply_button.count() == 0:
        apply_button = page.get_by_text("Apply", exact=True)

    apply_button.first.wait_for(state="visible", timeout=15000)
    apply_button.first.click(force=True)

    data_table_heading = page.get_by_text(re.compile(r"^Data\s*Table$", re.I)).first
    data_table_heading.wait_for(state="visible", timeout=30000)
    data_table_heading.scroll_into_view_if_needed()

    log("[Checker] Waiting for result table rows to populate...")

    try:
        page.wait_for_function(
            """
            () => {
                const headings = [...document.querySelectorAll('body *')]
                    .filter(el => el.textContent?.trim() === 'DataTable');

                if (!headings.length) return false;

                const heading = headings[0];

                // Search upward for a container containing a table.
                let node = heading;
                let table = null;

                for (let i = 0; i < 8 && node; i++, node = node.parentElement) {
                    table = node.querySelector('table');
                    if (table) break;
                }

                // Fallback to any visible table on page.
                if (!table) {
                    const tables = [...document.querySelectorAll('table')];
                    table = tables.find(t => {
                        const r = t.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    });
                }

                if (!table) return false;

                const rows = [...table.querySelectorAll('tbody tr')]
                    .filter(row => {
                        const r = row.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    });

                if (!rows.length) return false;

                // Require real row content, not an empty/loading shell.
                // Also accept explicit "No Data" rows to prevent infinite staring.
                const populatedRows = rows.filter(row => {
                    const txt = row.innerText.trim();
                    const hasLength = txt.length > 5;
                    const notLoading = !/loading|processing|please wait/i.test(txt);
                    const isExplicitNoData = /no data|no results/i.test(txt);
                    return (hasLength && notLoading) || isExplicitNoData;
                });

                return populatedRows.length > 0;
            }
            """,
            timeout=120000
        )

    except PlaywrightTimeoutError:
        screenshot(page, game_name, "datatable_rows_timeout")
        raise RuntimeError(
            "ByteBrew result table did not populate within 120 seconds."
        )

    # Wait for the rendered table to settle.
    # We require the visible row count to remain unchanged across checks.
    log("Data found. Waiting for table to finish rendering...")

    previous_count = None
    stable_checks = 0

    for _ in range(20):
        try:
            row_count = page.locator("table tbody tr:visible").count()
        except Exception:
            row_count = 0

        if row_count > 0 and row_count == previous_count:
            stable_checks += 1
        else:
            stable_checks = 0

        if stable_checks >= 3:
            break

        previous_count = row_count
        page.wait_for_timeout(750)

    # Extra small settle period for ByteBrew's menu/controls after table render.
    page.wait_for_timeout(1500)

    log("Funnel data fully loaded.")


def download_csv(page, output_path, game_name):
    """
    Open the three-dot menu on the result table and download the CSV.

    Works for both:
      - Funnel Explorer ("Data Table")
      - Mechanics ("DataTable")

    IMPORTANT:
    Mechanics shows a visible "Download CSV" menu item but does not necessarily
    expose the Funnel-specific #explorer-datatable-download element.
    Therefore we click the visible menu text first, and use the old ID only
    as a fallback.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log("Opening Data Table download menu...")

    data_table_heading = page.get_by_text(
        re.compile(r"^Data\s*Table$", re.I)
    ).first

    data_table_heading.wait_for(state="visible", timeout=15000)
    data_table_heading.scroll_into_view_if_needed()
    page.wait_for_timeout(500)

    heading_box = data_table_heading.bounding_box()

    if not heading_box:
        screenshot(page, game_name, "datatable_heading_no_box")
        raise RuntimeError("Could not determine Data Table heading position.")

    heading_y_center = heading_box["y"] + heading_box["height"] / 2
    heading_right = heading_box["x"] + heading_box["width"]

    # 1. Try direct icon / dropdown class selectors first
    menu_button = None
    icon_selectors = [
        ".fa-ellipsis-v", ".fa-ellipsis-h", ".dropdown-toggle",
        '[data-toggle="dropdown"]', ".fa-cog", ".fa-bars"
    ]
    for sel in icon_selectors:
        items = page.locator(sel)
        for i in range(items.count()):
            try:
                item = items.nth(i)
                if item.is_visible():
                    box = item.bounding_box()
                    if box and abs((box["y"] + box["height"]/2) - heading_y_center) <= 60:
                        menu_button = item
                        break
            except Exception:
                pass
        if menu_button:
            break

    # 2. Fallback: Search all DOM elements (including i, span, div tags) on the same header row
    if not menu_button:
        candidates = page.locator('button, [role="button"], a, [onclick], [tabindex="0"], i, span, div')
        nearby = []

        for i in range(candidates.count()):
            try:
                el = candidates.nth(i)
                if not el.is_visible():
                    continue

                box = el.bounding_box()
                if not box:
                    continue

                y_center = box["y"] + box["height"] / 2
                same_row = abs(y_center - heading_y_center) <= 50
                to_right = box["x"] > heading_right + 10
                reasonably_small = 5 <= box["width"] <= 180 and 5 <= box["height"] <= 100

                if same_row and to_right and reasonably_small:
                    nearby.append((box["x"], el, box))
            except Exception:
                pass

        if nearby:
            nearby.sort(key=lambda item: item[0])
            menu_button = nearby[-1][1]

    if not menu_button:
        screenshot(page, game_name, "datatable_header_controls_not_found")
        raise RuntimeError("Could not find controls on the Data Table header row.")

    log("Clicking Data Table rightmost header control...")
    menu_button.click()
    page.wait_for_timeout(400)

    # ------------------------------------------------------------
    # PRIMARY METHOD:
    # Mechanics clearly renders a visible "Download CSV" item.
    # Click that directly.
    # ------------------------------------------------------------
    visible_download = page.get_by_text(
        re.compile(r"^\s*Download\s+CSV\s*$", re.I)
    )

    download_item = None

    for i in range(visible_download.count()):
        try:
            item = visible_download.nth(i)
            if item.is_visible():
                download_item = item
                break
        except Exception:
            pass

    if download_item is not None:
        log('Visible "Download CSV" menu item found.')
        log("Downloading CSV...")

        with page.expect_download(timeout=30000) as download_info:
            download_item.click()

        download = download_info.value

    else:
        # --------------------------------------------------------
        # FALLBACK METHOD:
        # Funnel Explorer historically exposed this exact ID.
        # --------------------------------------------------------
        datatable_download = page.locator("#explorer-datatable-download")

        try:
            datatable_download.wait_for(state="visible", timeout=3000)
        except PlaywrightTimeoutError:
            screenshot(page, game_name, "download_csv_item_not_found")

            # Print the menu's visible text for debugging.
            try:
                visible_texts = page.locator("body").inner_text()
                lines = [
                    line.strip()
                    for line in visible_texts.splitlines()
                    if "download" in line.lower()
                ]
                log(f"Visible download-related text: {lines[:10]}")
            except Exception:
                pass

            raise RuntimeError(
                'Three-dot menu opened, but no visible "Download CSV" item '
                "could be found."
            )

        log("Using Funnel Explorer download control.")
        log("Downloading CSV...")

        with page.expect_download(timeout=30000) as download_info:
            datatable_download.click()

        download = download_info.value

    if output_path.exists():
        output_path.unlink()

    download.save_as(str(output_path))
    log(f"CSV saved: {output_path}")


def get_game_users_flexible(df, candidates):
    if 'EVENT' not in df.columns or 'USERS' not in df.columns:
        return 0
    matches = df[df['EVENT'].astype(str).isin(candidates)]
    if not matches.empty:
        return matches['USERS'].iloc[0]
    return 0


def get_ad_users_flexible(df_ad, patterns):
    if 'EVENT' not in df_ad.columns:
        return 0
    for pattern in patterns:
        event_row = df_ad[df_ad['EVENT'].astype(str).str.endswith(pattern, na=False)]
        if not event_row.empty and 'USERS' in event_row.columns:
            return event_row['USERS'].sum()
    return 0


def calculate_metrics_for_version(df_game, df_ad):
    metrics = {}

    # FAILSAFE: If ByteBrew exported as Line chart (has DATE col), filter to the most recent date ONLY
    if 'DATE' in df_game.columns and not df_game.empty:
        df_game = df_game[df_game['DATE'] == df_game['DATE'].max()].copy()
    if 'DATE' in df_ad.columns and not df_ad.empty:
        df_ad = df_ad[df_ad['DATE'] == df_ad['DATE'].max()].copy()

    def get_level_variations(prefix):
        return [f"{prefix} - levelStarted", f"{prefix} - level_started", f"{prefix} - level_start"]

    level_event_map = {
        'total_users': ['A - new_user', 'A - newUser', 'new_user'],
        20: get_level_variations('B'), 50: get_level_variations('C'), 70: get_level_variations('D'),
        100: get_level_variations('E'), 150: get_level_variations('F'), 200: get_level_variations('G'),
    }

    total_users = get_game_users_flexible(df_game, level_event_map['total_users'])
    metrics['Total Users'] = int(total_users) if total_users > 0 else 0

    for level in [20, 50, 70, 100, 150, 200]:
        users_at_level = get_game_users_flexible(df_game, level_event_map[level])
        metrics[f'% of users at {level}'] = users_at_level / total_users if total_users > 0 else 0.0

    for level in [10, 20, 40, 70, 100]:
        users_at_ad = get_ad_users_flexible(df_ad, [f"ads_{level}", f"adShown_{level}"])
        metrics[f'% of users at Ads {level}'] = users_at_ad / total_users if total_users > 0 else 0.0

    if total_users > 0 and 'EVENT' in df_ad.columns and 'EVENT AMOUNT' in df_ad.columns:
        inter_mask = df_ad['EVENT'].astype(str).str.contains('J', na=False) & df_ad['EVENT'].astype(str).str.contains('inter', na=False)
        reward_mask = df_ad['EVENT'].astype(str).str.contains('undefined', na=False) & df_ad['EVENT'].astype(str).str.contains('reward', na=False)
        total_ad_events = df_ad.loc[inter_mask | reward_mask, 'EVENT AMOUNT'].sum()
        metrics['Avg Ad per user'] = total_ad_events / total_users
    else:
        metrics['Avg Ad per user'] = 0.0

    return metrics


def process_dp1_merge(output_dir, game):
    """Merges downloaded PartA and PartB CSVs into <Output Name>.xlsx."""
    game_name = game["game_name"]
    output_name = game.get("output_name") or game_name

    part_a_files = list(output_dir.glob(f"{safe_name(game_name)}_*PartA*.csv"))
    part_b_files = list(output_dir.glob(f"{safe_name(game_name)}_*PartB*.csv"))

    if not part_a_files or not part_b_files:
        all_csvs = list(output_dir.glob(f"{safe_name(game_name)}_*.csv"))
        for f in all_csvs:
            try:
                df = pd.read_csv(f)
                if 'EVENT' in df.columns:
                    if df['EVENT'].astype(str).str.contains('A - new_user', na=False).any():
                        part_a_files = [f]
                    elif df['EVENT'].astype(str).str.contains('A - adShown_10', na=False).any():
                        part_b_files = [f]
            except Exception:
                pass

    if not part_a_files or not part_b_files:
        log(f"[{game_name}] PartA or PartB CSV missing; skipping DP1 Excel report.")
        return None

    log(f'[{game_name}] Generating final DP1 report: "{output_name}.xlsx"...')
    
    # Load raw CSVs
    df_ret = pd.read_csv(part_a_files[0])
    df_ad = pd.read_csv(part_b_files[0])

    # Unify Line and Bar chart formats: Drop DATE column if present so structure strictly matches 6 standard columns
    if 'DATE' in df_ret.columns:
        df_ret = df_ret.drop(columns=['DATE'])
    if 'DATE' in df_ad.columns:
        df_ad = df_ad.drop(columns=['DATE'])

    df_ret.to_csv(part_a_files[0], index=False)
    df_ad.to_csv(part_b_files[0], index=False)

    metrics = calculate_metrics_for_version(df_ret, df_ad)

    kpi_order = [
        'Total Users', '% of users at 20', '% of users at 50', '% of users at 70',
        '% of users at 100', '% of users at 150', '% of users at 200',
        '% of users at Ads 10', '% of users at Ads 20', '% of users at Ads 40',
        '% of users at Ads 70', '% of users at Ads 100', 'Avg Ad per user'
    ]

    report_data = {
        'KPI': kpi_order,
        game_name: [metrics.get(kpi, 0.0) for kpi in kpi_order]
    }

    final_report = pd.DataFrame(report_data)
    excel_path = output_dir / f"{safe_name(output_name)}.xlsx"

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        final_report.to_excel(writer, sheet_name='DP1_Report', index=False)
        sheet = writer.sheets['DP1_Report']

        percent_format = '0.00%'
        float_format = '0.00'
        int_format = '0'
        center_align = Alignment(horizontal='center', vertical='center')
        dark_red_font = Font(color='632523', bold=False)
        bold_font = Font(bold=True)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for column in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = max_length + 4

        for row in sheet.iter_rows(min_row=2):
            kpi_name = str(row[0].value or '')
            font_to_apply = Font(bold=False)
            if '%' in kpi_name and 'Ads' in kpi_name:
                font_to_apply = dark_red_font
            elif kpi_name == 'Avg Ad per user':
                font_to_apply = bold_font

            for cell in row:
                cell.alignment = center_align
                cell.font = font_to_apply

                if cell.column == 1:
                    continue

                if kpi_name == 'Total Users':
                    cell.number_format = int_format
                elif '%' in kpi_name:
                    cell.number_format = percent_format
                elif kpi_name == 'Avg Ad per user':
                    cell.number_format = float_format

    log(f"[{game_name}] Final report created: {excel_path.name}")

    # Requirement 1: Instantly delete the raw PartA and PartB CSVs once the Excel report completes successfully
    for csv_file in set(part_a_files + part_b_files):
        try:
            if csv_file.exists():
                csv_file.unlink()
                log(f"[{game_name}] Cleaned up raw CSV: {csv_file.name}")
        except Exception as err:
            log(f"[{game_name}] Failed to delete {csv_file.name}: {err}")

    return excel_path
def process_game(base_page, game, email=None, password=None):
    """
    Process every configured saved funnel sequentially.
    Tabs are nuked and recreated for each funnel to guarantee zero state bleeding.
    """
    global CURRENT_DATE_DIR
    game_name = game["game_name"]
    saved_funnels = game.get("saved_funnels", [])

    if not saved_funnels:
        raise RuntimeError(
            f'No Saved Funnel 1 / Saved Funnel 2 configured for "{game_name}".'
        )

    # Establish the daily target directory immediately so logs and screenshots route correctly from step 1
    folder_date = output_date_for_preset(game["date_range"])
    output_dir = UPLOAD_DIR / folder_date.isoformat()
    output_dir.mkdir(parents=True, exist_ok=True)
    CURRENT_DATE_DIR = output_dir

    outputs = []
    browser_context = base_page.context

    # Close the initial blank page passed from app.py to keep things completely isolated
    try:
        base_page.close()
    except Exception:
        pass

    for funnel_index, saved_funnel in enumerate(saved_funnels, start=1):
        print()
        log(
            f'[{game_name}] Funnel {funnel_index}/{len(saved_funnels)}: '
            f'"{saved_funnel}"'
        )

        # Nuke previous state: open a completely fresh tab for this funnel
        page = browser_context.new_page()
        page.set_default_timeout(15000)

        try:
            # Fresh tab login and navigation
            open_game(page, game, email=email, password=password)

            # 1. Load saved preset first
            load_saved_funnel(page, saved_funnel, game_name)

            # 2. Force Top Main Date Range AFTER loading preset
            set_date_range(page, game["date_range"], game_name)

            # 3. Apply Mechanic User Install Date cohort
            if "mechanic" in game["page"].lower():
                configure_install_date_filter(
                    page,
                    game["date_range"],
                    game_name
                )
                
                # QQ4 Killshot Check
                verify_filter_pills(page, game_name)

            # 4. If Build Version or Country is configured in Sheet, ensure selection without unchecking
            if game.get("build_version") or game.get("country"):
                configure_build_and_geo(page, game.get("build_version"), game.get("country"), game_name)

            # 5. Apply funnel (Ignore UI Graph mode; raw data formatting handled in Python)
            apply_funnel(page, game_name)

            # Save individual CSVs with Game Name + Saved Funnel
            output_file = output_dir / (
                f'{safe_name(game["game_name"])}_'
                f'{safe_name(saved_funnel)}.csv'
            )

            download_csv(page, output_file, game_name)
            outputs.append(output_file)

            log(f'Completed "{saved_funnel}".')
        finally:
            # Tab Nuke: Close the tab so the next Part/Funnel starts 100% fresh
            try:
                page.close()
            except Exception:
                pass

    # Automatically generate merged DP1 Excel report (<Output Name>.xlsx)
    merged_report = process_dp1_merge(output_dir, game)
    if merged_report:
        outputs.append(merged_report)

    # Requirement 3: Success Cleanup - If the code reaches this line, no exceptions were thrown. Delete debug screenshots.
    for png in output_dir.glob(f"{safe_name(game_name)}_*.png"):
        try:
            png.unlink()
            log(f"[{game_name}] Cleaned up temporary debug screenshot: {png.name}")
        except Exception:
            pass

    return outputs


def main():
    games = load_games()

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    print()
    print("=" * 74)
    print("GAMES DP1 DATA - BYTEBREW DOWNLOAD")
    print("=" * 74)
    print(f"Active games: {len(games)}")
    print()

    successes = []
    failures = []

    with sync_playwright() as p:
        context = None
        for channel in ["chrome", "msedge", None]:
            try:
                kwargs = {
                    "user_data_dir": str(PROFILE_DIR),
                    "headless": False,
                    "accept_downloads": True,
                    "viewport": {"width": 1600, "height": 950},
                    "args": ["--start-maximized"],
                }
                if channel:
                    kwargs["channel"] = channel
                context = p.chromium.launch_persistent_context(**kwargs)
                log(f"Browser launched using: {channel or 'bundled chromium'}")
                break
            except Exception:
                continue

        if not context:
            raise RuntimeError("Could not launch system or bundled browser.")

        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(15000)

        for index, game in enumerate(games, start=1):
            print()
            print("-" * 74)
            print(f'[{index}/{len(games)}] {game["game_name"]}')
            print("-" * 74)

            try:
                outputs = process_game(page, game)
                successes.append((game["game_name"], outputs))
                print(f'✓ {game["game_name"]} complete ({len(outputs)} file(s))')
            except Exception as exc:
                failures.append((game["game_name"], str(exc)))
                screenshot(page, game["game_name"], "error")
                print(f'✗ {game["game_name"]} FAILED: {exc}')
                traceback.print_exc()

                # Continue to the next game instead of stopping all downloads.
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(500)
                except Exception:
                    pass

        context.close()

    print()
    print("=" * 74)
    print("RUN COMPLETE")
    print("=" * 74)
    print(f"Successful: {len(successes)}")
    print(f"Failed:     {len(failures)}")

    for game_name, outputs in successes:
        print(f"  ✓ {game_name}:")
        for output in outputs:
            print(f"      {output}")

    for game_name, error in failures:
        print(f"  ✗ {game_name}: {error}")

    # Write execution summary inside System Files logs
    if games:
        summary_dir = SYSTEM_DIR / "logs"
        summary_dir.mkdir(parents=True, exist_ok=True)
        summary_file = summary_dir / "run_summary.txt"

        with summary_file.open("w", encoding="utf-8") as f:
            f.write(f"Games DP1 Data run: {datetime.now().isoformat(timespec='seconds')}\n")
            f.write(f"Successful: {len(successes)}\n")
            f.write(f"Failed: {len(failures)}\n\n")

            for game_name, outputs in successes:
                for output in outputs:
                    f.write(f"SUCCESS | {game_name} | {output.name}\n")

            for game_name, error in failures:
                f.write(f"FAILED | {game_name} | {error}\n")

    print()
    if failures:
        print("Some games failed. Check the debug folder and run_summary.txt.")
    else:
        print("All active games downloaded successfully.")


if __name__ == "__main__":
    main()
